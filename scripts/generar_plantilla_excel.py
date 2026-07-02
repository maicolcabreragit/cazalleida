#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera una plantilla de Excel vacía y profesional con 50 casillas.
Incluye fórmulas dinámicas de Excel para que los totales se calculen solos.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Paleta de colores ────────────────────────────────────────────────────────
C_VERDE_OSC  = "1F3D2B"
C_VERDE_MED  = "2F5A3F"
C_DORADO     = "A9852F"
C_CREMA      = "F7F3E8"
C_CREMA2     = "EDE8D5"
C_BLANCO     = "FFFFFF"

C_HDR_BG     = "1F3D2B"   # Cabecera verde oscuro
C_HDR_TX     = "FFFFFF"

# ─── Estilos helper ───────────────────────────────────────────────────────────
def borde(color="CCCCCC", grosor="thin"):
    s = Side(border_style=grosor, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(color):
    return PatternFill("solid", fgColor=color)

def font(bold=False, color="000000", size=10, italic=False, name="Calibri"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─── Crear libro ──────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Plantilla 50 Cazadores"

# Habilitar líneas de cuadrícula visibles
ws.views.sheetView[0].showGridLines = True

# Anchos de columna: Nº | NOMBRE | DNI | NUM. TEL. | COTO | PAGO | IMPORTE
anchos = [6, 44, 17, 16, 15, 12, 12]
for i, w in enumerate(anchos, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ─── Fila 1: Título principal ─────────────────────────────────────────────────
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "REGISTRO Y CONTROL DE CAZADORES  ·  PLANTILLA VACÍA"
c.font  = Font(name="Calibri", bold=True, color=C_HDR_TX, size=13)
c.fill  = fill(C_HDR_BG)
c.alignment = align("center")
ws.row_dimensions[1].height = 28

# ─── Fila 2: Instrucciones / Subtítulo ────────────────────────────────────────
ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "Rellene los datos de los cazadores. Los totales inferiores se calcularán automáticamente."
c.font  = Font(name="Calibri", italic=True, color=C_DORADO, size=9.5)
c.fill  = fill(C_CREMA)
c.alignment = align("center")
ws.row_dimensions[2].height = 18

# ─── Fila 3: Cabeceras ────────────────────────────────────────────────────────
cabeceras = ["Nº", "NOMBRE Y APELLIDOS", "DNI", "NUM. TEL.", "COTO", "PAGO", "IMPORTE"]
alineacion_cab = ["center", "left", "center", "center", "center", "center", "center"]
for col, (cab, alin) in enumerate(zip(cabeceras, alineacion_cab), 1):
    c = ws.cell(row=3, column=col, value=cab)
    c.font      = Font(name="Calibri", bold=True, color=C_HDR_TX, size=10)
    c.fill      = fill(C_HDR_BG)
    c.alignment = align(alin)
    c.border    = borde("3D7055", "medium")
ws.row_dimensions[3].height = 22

# ─── Validación de datos para la columna PAGO (dropdown list) ─────────────────
dv_pago = DataValidation(type="list", formula1='"Pagado,Falta,S.C"', allow_blank=True)
dv_pago.error ='El valor introducido no es válido. Elija de la lista: Pagado, Falta o S.C'
dv_pago.errorTitle = 'Valor incorrecto'
dv_pago.prompt = 'Seleccione: Pagado, Falta o S.C'
dv_pago.promptTitle = 'Estado de Pago'
ws.add_data_validation(dv_pago)

# ─── Agregar 50 filas vacías con estilo ───────────────────────────────────────
rango_inicio = 4
rango_fin = 53

for fila in range(rango_inicio, rango_fin + 1):
    # Alternar color crema muy suave de fondo para lectura cómoda
    bg_fila = C_CREMA if fila % 2 == 0 else C_BLANCO
    row_fill = fill(bg_fila)
    borde_fino = borde("CCCCCC", "thin")

    # Nº
    c = ws.cell(row=fila, column=1, value=fila - rango_inicio + 1)
    c.fill = row_fill; c.border = borde_fino
    c.font = font(color="888888", size=9)
    c.alignment = align("center")

    # Nombre (vacío)
    c = ws.cell(row=fila, column=2, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(size=10)
    c.alignment = align("left")

    # DNI (vacío)
    c = ws.cell(row=fila, column=3, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(size=10)
    c.alignment = align("center")

    # Teléfono (vacío)
    c = ws.cell(row=fila, column=4, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(size=10)
    c.alignment = align("center")

    # Coto (puedes escribir o dejar vacío, ponemos "Grayena" como sugerencia por defecto)
    c = ws.cell(row=fila, column=5, value="Grayena")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(size=10)
    c.alignment = align("center")

    # Pago (vacío)
    c = ws.cell(row=fila, column=6, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(bold=True, size=10)
    c.alignment = align("center")
    dv_pago.add(c) # Agregar dropdown

    # Importe (vacío)
    c = ws.cell(row=fila, column=7, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(bold=True, size=10)
    c.alignment = align("right")
    # Formato numérico de moneda: 0 €
    c.number_format = '#,##0" €"'

    ws.row_dimensions[fila].height = 19

# ─── Fila vacía separadora ────────────────────────────────────────────────────
fila_sep = rango_fin + 1
ws.row_dimensions[fila_sep].height = 8
fila_res = fila_sep + 1

# ─── Tabla de Resumen con Fórmulas Dinámicas de Excel ─────────────────────────
ws.merge_cells(f"A{fila_res}:G{fila_res}")
c = ws.cell(row=fila_res, column=1, value="RESUMEN AUTOMÁTICO DE DATOS (FÓRMULAS EXCEL)")
c.font = Font(name="Calibri", bold=True, color=C_HDR_TX, size=11)
c.fill = fill(C_HDR_BG)
c.alignment = align("center")
ws.row_dimensions[fila_res].height = 22
fila_res += 1

# Definir las fórmulas
rango_nombres = f"B{rango_inicio}:B{rango_fin}"
rango_pagos   = f"F{rango_inicio}:F{rango_fin}"
rango_importes = f"G{rango_inicio}:G{rango_fin}"

resumen_formulas = [
    ("Total de cazadores registrados (con nombre)",   f"=COUNTA({rango_nombres})",        C_VERDE_OSC, C_CREMA),
    ("✓  Pagados",                                     f'=COUNTIF({rango_pagos}, "Pagado")', C_VERDE_MED, "E0F5E0"),
    ("✗  Pendientes de pago (Falta)",                  f'=COUNTIF({rango_pagos}, "Falta")',  "B22222",    "FFE0E0"),
    ("·  Sin Coste (S.C)",                             f'=COUNTIF({rango_pagos}, "S.C")',    "1A4E8C",    "E0ECFF"),
    ("€  Total Recaudado (Cobrado)",                   f'=SUMIF({rango_pagos}, "Pagado", {rango_importes})', C_VERDE_OSC, "E0F5E0"),
    ("€  Total Pendiente (Por Cobrar)",                f'=SUMIF({rango_pagos}, "Falta", {rango_importes})', "B22222",    "FFE0E0"),
    ("€  Total General (Proyectado)",                  f"=SUM({rango_importes})",          C_VERDE_OSC, C_CREMA2),
]

for label, formula, tx_col, bg_col in resumen_formulas:
    # Etiqueta
    ws.merge_cells(f"A{fila_res}:E{fila_res}")
    c = ws.cell(row=fila_res, column=1, value=label)
    c.font = Font(name="Calibri", bold=True, color=tx_col, size=10)
    c.fill = fill(bg_col)
    c.alignment = align("right")
    c.border = borde("BBBBBB")

    # Fórmula
    ws.merge_cells(f"F{fila_res}:G{fila_res}")
    c = ws.cell(row=fila_res, column=6, value=formula)
    c.font = Font(name="Calibri", bold=True, color=tx_col, size=11)
    c.fill = fill(bg_col)
    c.alignment = align("center")
    c.border = borde("BBBBBB")
    
    # Si es total en dinero, formatearlo
    if "SUM" in formula:
        c.number_format = '#,##0" €"'

    ws.row_dimensions[fila_res].height = 19
    fila_res += 1

# ─── Bloqueo de paneles y filtros ─────────────────────────────────────────────
ws.freeze_panes = "A4"

# Guardar plantilla
output_file = r"C:\cazalleida . com\cazadores-PLANTILLA-50-cazadores.xlsx"
wb.save(output_file)
print(f"OK Plantilla vacia guardada: {output_file}")
