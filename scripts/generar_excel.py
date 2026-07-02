#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera la tabla Excel profesional de cazadores del Coto Grayena.
Columnas: Nº | NOMBRE | DNI | NUM. TEL. | COTO | PAGO | IMPORTE
Incluye 10 filas adicionales vacías para añadir cazadores fácilmente.
"""

import unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Paleta de colores ────────────────────────────────────────────────────────
C_VERDE_OSC  = "1F3D2B"
C_VERDE_MED  = "2F5A3F"
C_DORADO     = "A9852F"
C_CREMA      = "F7F3E8"
C_CREMA2     = "EDE8D5"
C_BLANCO     = "FFFFFF"

C_FALTA_BG   = "FFE0E0"
C_FALTA_TX   = "B22222"
C_SC_BG      = "E0ECFF"
C_SC_TX      = "1A4E8C"
C_PAG_BG     = "E0F5E0"
C_PAG_TX     = "1A6B1A"
C_GRUPO_BG   = "FFFAE8"   # fondo miembro de grupo
C_HDR_BG     = "1F3D2B"   # cabecera
C_HDR_TX     = "FFFFFF"

# ─── Datos ────────────────────────────────────────────────────────────────────
# (nombre, dni, tel, pago, importe, es_lider)
# pago:    "Pagado" | "Falta" | "S.C"
# importe: "300€"  | "500€" | "600€" | "S.C" | ""

GRUPOS = [
    # Benito Gili Florensa (presidente / coordinador)
    [("Benito Gili Florensa",                       "40898319H",    "699 94 97 38",  "S.C",    "S.C",  True)],

    # Enric Cabré y grupo (4 personas · 300 €)
    [("Enric Cabré Solá",                           "40834218H",    "610 257 117",   "Pagado", "300€", True),
     ("Antonio Sunalla Raimat",                     "40838458A",    "",              "Pagado", "300€", False),
     ("Jordi Sans Pascuet",                         "40841054T",    "",              "Pagado", "300€", False),
     ("Josep Cots Pelegri",                         "40847530J",    "",              "Pagado", "300€", False)],

    # Emilio Useno Castno (solo · 600 €)
    [("Emilio Useno Castno",                        "37323935H",    "696 624 264",   "Pagado", "600€", True)],

    # Ferran Galcerà Solé (solo · S.C)
    [("Ferran Galcerà Solé",                        "47936701-D",   "636 451 78",    "S.C",    "S.C",  True)],

    # Francesc Sánchez Besora (solo · Falta · 600€ esperado)
    [("Francesc Sánchez Besora",                    "46723855E",    "627 476 830",   "Falta",  "600€", True)],

    # Francisco Zarza y grupo (3 personas · 600 €)
    [("Francisco Zarza Garcia",                     "39699060W",    "674 918 163",   "Pagado", "600€", True),
     ("Victor Manuel Zarza Garcia",                 "39699767L",    "",              "Pagado", "600€", False),
     ("Yeray Zarza Callejón",                       "49316079P",    "",              "Pagado", "600€", False)],

    # Diego Jose Galera Quisano (solo · 500 €)
    [("Diego Jose Galera Quisano",                  "39473017-A",   "617 731 219",   "Pagado", "500€", True)],

    # Joaquin Feixo Fannagoldo (solo · 600 €· NIE/And)
    [("Joaquin Feixo Fannagoldo",                   "06643116 (And)","604 119 191",  "Pagado", "600€", True)],

    # Jordi Vidal García + Jorge Vidal Miró (2 personas · 300 € · Falta)
    [("Jordi Vidal García",                         "39938597V",    "660 964 600",   "Falta",  "300€", True),
     ("Jorge Vidal Miró",                           "39874279F",    "",              "Falta",  "300€", False)],

    # José A. Galera Gea (solo · 500 €)
    [("José A. Galera Gea",                         "39728517L",    "722 222 795",   "Pagado", "500€", True)],

    # Jose A. Rodriguez Sanchez (solo · Falta · 600€ esperado)
    [("Jose A. Rodriguez Sanchez",                  "43499011P",    "689 421 078",   "Falta",  "600€", True)],

    # Jose Bernal Marnenda + Abel Bernal (2 personas)
    [("Jose Bernal Marnenda",                       "43500907-H",   "600 321 681",   "Pagado", "600€", True),
     ("Abel Bernal Galvez",                         "53033535C",    "",              "Falta",  "600€", False)],

    # Josep María Reves (solo · S.C)
    [("Josep María Reves",                          "78050679H",    "618 295 499",   "S.C",    "S.C",  True)],

    # Juan Antonio Fdez. Suarez y grupo (3 personas · 500 €)
    [("Juan Antonio Fdez. Suarez",                  "39675622R",    "676 847 710",   "Pagado", "500€", True),
     ("Sebastian Hervás Soriano",                   "39854673C",    "",              "Pagado", "500€", False),
     ("Adrian Fdez Izquierdo",                      "49647703H",    "",              "Pagado", "500€", False)],

    # Juan Corominas de Dios + Inga Ermolova (2 personas · 600 €)
    [("Juan Corominas de Dios",                     "77260203Y",    "606 449 319",   "Pagado", "600€", True),
     ("Inga Ermolova",                              "47918378V",    "",              "Pagado", "600€", False)],

    # Juan Francisco de la Cruz Sanchez Plana (solo · S.C)
    [("Juan Francisco de la Cruz Sanchez Plana",    "40876029S",    "686 943 349",   "S.C",    "S.C",  True)],

    # Llorenç Ruiz Sala (solo · 600 €)
    [("Llorenç Ruiz Sala",                          "34758374S",    "638 387 588",   "Pagado", "600€", True)],

    # Manuel Castillo Lopez + Carlos Castillo Leon (2 personas)
    [("Manuel Castillo Lopez",                      "4498318P",     "628 504 033",   "Pagado", "600€", True),
     ("Carlos Castillo Leon",                       "77795483F",    "",              "Falta",  "600€", False)],

    # Miquel Angel Rodriguez y grupo Jaldo (4 personas · 500 €)
    [("Miquel Angel Rodriguez",                     "46719456Q",    "692 589 875",   "Pagado", "500€", True),
     ("Miquel Jaldo Pelegrino",                     "46518981D",    "",              "Pagado", "500€", False),
     ("Miquel Jaldo Delpino",                       "49288378E",    "",              "Pagado", "500€", False),
     ("Jesus Jaldo Pelegrina",                      "46537187E",    "",              "Pagado", "500€", False)],

    # Miquel Galcerà Altés (solo · S.C)
    [("Miquel Galcerà Altés",                       "78575149-L",   "",              "S.C",    "S.C",  True)],

    # Ramón Vidal Sabaté (solo · S.C)
    [("Ramón Vidal Sabaté",                         "40880798T",    "680 702 577",   "S.C",    "S.C",  True)],

    # Sebastian Altemir Riba (solo · Falta · 600€ esperado)
    [("Sebastian Altemir Riba",                     "53332504N",    "",              "Falta",  "600€", True)],

    # Sebastia Ros Barbera + Adriá Ros Pujol (2 personas · Falta · 600€ esperado)
    [("Sebastia Ros Barbera",                       "40864924L",    "625 146 666",   "Falta",  "600€", True),
     ("Adriá Ros Pujol",                            "78089599E",    "",              "Falta",  "600€", False)],

    # Sergio Valera Guerrero y grupo (3 personas · Falta · 600€ esperado)
    [("Sergio Valera Guerrero",                     "39888316Z",    "652 916 065",   "Falta",  "600€", True),
     ("Iker Valera Gisbert",                        "39940528Q",    "",              "Falta",  "600€", False),
     ("Juan Varela Ortiz",                          "25912415V",    "",              "Falta",  "600€", False)],
]

def normalizar(s):
    """Elimina acentos para ordenar correctamente."""
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()

# Ordenar grupos alfabéticamente por nombre del líder
GRUPOS.sort(key=lambda g: normalizar(g[0][0]))

# ─── Estilos helper ───────────────────────────────────────────────────────────
def borde(color="BBBBBB", grosor="thin"):
    s = Side(border_style=grosor, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(color):
    return PatternFill("solid", fgColor=color)

def font(bold=False, color="000000", size=10, italic=False, name="Calibri"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─── Crear libro ──────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Cazadores Grayena 2026"

# Mostrar líneas de cuadrícula
ws.views.sheetView[0].showGridLines = True

# Anchos de columna: Nº | NOMBRE | DNI | TEL | COTO | PAGO | IMPORTE
anchos = [6, 44, 17, 16, 13, 11, 12]
for i, w in enumerate(anchos, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ─── Fila 1: Título principal ─────────────────────────────────────────────────
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "COTO GRAYENA DE LES GARRIGUES  ·  A.P.C L-10.199  ·  JULIO 2026"
c.font  = Font(name="Calibri", bold=True, color=C_HDR_TX, size=14)
c.fill  = fill(C_HDR_BG)
c.alignment = align("center")
ws.row_dimensions[1].height = 30

# ─── Fila 2: Subtítulo ────────────────────────────────────────────────────────
ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "AUTORIZACIÓN ESPECIAL DE CAZA DEL CONEJO POR DAÑOS A LA AGRICULTURA"
c.font  = Font(name="Calibri", bold=True, color=C_DORADO, size=10)
c.fill  = fill(C_CREMA)
c.alignment = align("center")
ws.row_dimensions[2].height = 18

# ─── Fila 3: Cabeceras ────────────────────────────────────────────────────────
cabeceras = ["Nº", "NOMBRE Y APELLIDOS", "DNI", "NUM. TEL.", "COTO", "PAGO", "IMPORTE"]
alineacion_cab = ["center", "left", "center", "center", "center", "center", "center"]
for col, (cab, alin) in enumerate(zip(cabeceras, alineacion_cab), 1):
    c = ws.cell(row=3, column=col, value=cab)
    c.font      = Font(name="Calibri", bold=True, color=C_HDR_TX, size=10)
    c.fill      = fill(C_HDR_BG)
    c.alignment = align(alin)
    c.border    = borde("3D7055", "medium")
ws.row_dimensions[3].height = 22

# ─── Validación de datos para la columna PAGO (dropdown list) ─────────────────
dv_pago = DataValidation(type="list", formula1='"Pagado,Falta,S.C"', allow_blank=True)
dv_pago.error ='El valor introducido no es válido. Elija de la lista: Pagado, Falta o S.C'
dv_pago.errorTitle = 'Valor incorrecto'
dv_pago.prompt = 'Seleccione: Pagado, Falta o S.C'
dv_pago.promptTitle = 'Estado de Pago'
ws.add_data_validation(dv_pago)

# ─── Filas de datos (Cazadores Existentes) ────────────────────────────────────
fila = 4
num  = 1
alt  = True   # alterna fondo en filas de líderes solos

for grupo in GRUPOS:
    for idx, (nombre, dni, tel, pago, importe, es_lider) in enumerate(grupo):
        es_grupo_miembro = not es_lider

        # ── color de fondo ──
        if pago == "Falta":
            bg = C_FALTA_BG
        elif pago == "S.C":
            bg = C_SC_BG
        elif es_grupo_miembro:
            bg = C_GRUPO_BG
        else:
            bg = C_CREMA if alt else C_BLANCO

        row_fill   = fill(bg)
        borde_fino = borde("CCCCCC", "thin")

        # ── nº ──
        c = ws.cell(row=fila, column=1, value=num)
        c.fill = row_fill; c.border = borde_fino
        c.font = font(color="777777", size=9)
        c.alignment = align("center")

        # ── nombre ──
        if es_grupo_miembro:
            txt = "    " + nombre
            c = ws.cell(row=fila, column=2, value=txt)
            c.font = Font(name="Calibri", italic=True, color="555555", size=10)
        else:
            c = ws.cell(row=fila, column=2, value=nombre)
            c.font = Font(name="Calibri", bold=(len(grupo) > 1), color="111111", size=10)
        c.fill = row_fill; c.border = borde_fino
        c.alignment = align("left")

        # ── DNI ──
        c = ws.cell(row=fila, column=3, value=dni)
        c.fill = row_fill; c.border = borde_fino
        c.font = font(color="222222", size=10)
        c.alignment = align("center")

        # ── Tel ──
        c = ws.cell(row=fila, column=4, value=tel)
        c.fill = row_fill; c.border = borde_fino
        c.font = font(color="1A1A6B", size=10)
        c.alignment = align("center")

        # ── Coto ──
        c = ws.cell(row=fila, column=5, value="Grayena")
        c.fill = row_fill; c.border = borde_fino
        c.font = Font(name="Calibri", bold=True, color=C_VERDE_OSC, size=10)
        c.alignment = align("center")

        # ── PAGO ──
        c = ws.cell(row=fila, column=6, value=pago)
        c.fill = row_fill; c.border = borde_fino
        c.alignment = align("center")
        dv_pago.add(c) # dropdown
        if pago == "Falta":
            c.font = Font(name="Calibri", bold=True, color=C_FALTA_TX, size=10)
        elif pago == "S.C":
            c.font = Font(name="Calibri", bold=True, color=C_SC_TX, size=10)
        else:
            c.font = Font(name="Calibri", bold=True, color=C_PAG_TX, size=10)

        # ── IMPORTE (Formato numérico para fórmulas) ──
        val_imp = None
        if importe.endswith("€"):
            val_imp = int(importe.replace("€", ""))
        elif importe == "S.C":
            val_imp = "S.C"
        else:
            val_imp = ""
            
        c = ws.cell(row=fila, column=7, value=val_imp)
        c.fill = row_fill; c.border = borde_fino
        c.alignment = align("right" if isinstance(val_imp, int) else "center")
        
        if isinstance(val_imp, int):
            c.number_format = '#,##0" €"'
            if pago == "Falta":
                c.font = Font(name="Calibri", bold=True, color=C_FALTA_TX, size=10)
            else:
                c.font = Font(name="Calibri", bold=True, color=C_VERDE_OSC, size=10)
        else:
            c.font = Font(name="Calibri", bold=True, color=C_SC_TX, size=10)

        ws.row_dimensions[fila].height = 17
        fila += 1
        num  += 1

    # Línea separadora gruesa al final de cada grupo
    for col in range(1, 8):
        cell = ws.cell(row=fila - 1, column=col)
        existing = cell.border
        cell.border = Border(
            left   = existing.left,
            right  = existing.right,
            top    = existing.top,
            bottom = Side(border_style="medium", color="AAAAAA")
        )

    alt = not alt

# ─── Escribir 10 filas adicionales vacías ─────────────────────────────────────
rango_vacias = 10
for i in range(rango_vacias):
    bg_fila = C_CREMA if num % 2 == 0 else C_BLANCO
    row_fill = fill(bg_fila)
    borde_fino = borde("CCCCCC", "thin")

    # Nº
    c = ws.cell(row=fila, column=1, value=num)
    c.fill = row_fill; c.border = borde_fino
    c.font = font(color="888888", size=9)
    c.alignment = align("center")

    # Nombre
    c = ws.cell(row=fila, column=2, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(size=10)
    c.alignment = align("left")

    # DNI
    c = ws.cell(row=fila, column=3, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(size=10)
    c.alignment = align("center")

    # Tel
    c = ws.cell(row=fila, column=4, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(size=10)
    c.alignment = align("center")

    # Coto
    c = ws.cell(row=fila, column=5, value="Grayena")
    c.fill = row_fill; c.border = borde_fino
    c.font = Font(name="Calibri", bold=True, color=C_VERDE_OSC, size=10)
    c.alignment = align("center")

    # Pago
    c = ws.cell(row=fila, column=6, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(bold=True, size=10)
    c.alignment = align("center")
    dv_pago.add(c) # dropdown

    # Importe
    c = ws.cell(row=fila, column=7, value="")
    c.fill = row_fill; c.border = borde_fino
    c.font = font(bold=True, size=10)
    c.alignment = align("right")
    c.number_format = '#,##0" €"'

    ws.row_dimensions[fila].height = 17
    fila += 1
    num  += 1

# ─── Fila vacía antes del resumen ─────────────────────────────────────────────
ws.row_dimensions[fila].height = 8
fila += 1

# ─── Resumen de pagos con Fórmulas Dinámicas ──────────────────────────────────
ws.merge_cells(f"A{fila}:G{fila}")
c = ws.cell(row=fila, column=1, value="RESUMEN AUTOMÁTICO DE PAGOS (FÓRMULAS EXCEL)")
c.font = Font(name="Calibri", bold=True, color=C_HDR_TX, size=11)
c.fill = fill(C_HDR_BG)
c.alignment = align("center")
ws.row_dimensions[fila].height = 22
fila += 1

# Fórmulas basadas en el rango dinámico de las celdas (fila 4 a fila_fin)
fila_fin = fila - 3
rango_nombres = f"B4:B{fila_fin}"
rango_pagos   = f"F4:F{fila_fin}"
rango_importes = f"G4:G{fila_fin}"

resumen = [
    ("Total de cazadores registrados (con nombre)",   f"=COUNTA({rango_nombres})",        C_VERDE_OSC, C_CREMA),
    ("✓  Pagados",                                     f'=COUNTIF({rango_pagos}, "Pagado")', C_PAG_TX,    C_PAG_BG),
    ("✗  Pendientes de pago (Falta)",                  f'=COUNTIF({rango_pagos}, "Falta")',  C_FALTA_TX,  C_FALTA_BG),
    ("·  Sin Coste (S.C)",                             f'=COUNTIF({rango_pagos}, "S.C")',    C_SC_TX,     C_SC_BG),
    ("€  Total Recaudado (Cobrado)",                   f'=SUMIF({rango_pagos}, "Pagado", {rango_importes})', C_VERDE_OSC, C_PAG_BG),
    ("€  Total Pendiente (Por Cobrar)",                f'=SUMIF({rango_pagos}, "Falta", {rango_importes})', C_FALTA_TX,  C_FALTA_BG),
    ("€  Total General (Proyectado)",                  f"=SUM({rango_importes})",          C_VERDE_OSC, C_CREMA2),
]

for label, formula, tx_col, bg_col in resumen:
    ws.merge_cells(f"A{fila}:E{fila}")
    c = ws.cell(row=fila, column=1, value=label)
    c.font = Font(name="Calibri", bold=True, color=tx_col, size=10)
    c.fill = fill(bg_col)
    c.alignment = align("right")
    c.border = borde("BBBBBB")

    ws.merge_cells(f"F{fila}:G{fila}")
    c = ws.cell(row=fila, column=6, value=formula)
    c.font = Font(name="Calibri", bold=True, color=tx_col, size=11)
    c.fill = fill(bg_col)
    c.alignment = align("center")
    c.border = borde("BBBBBB")
    
    if "SUM" in formula:
        c.number_format = '#,##0" €"'

    ws.row_dimensions[fila].height = 18
    fila += 1

# ─── Ajustes finales ──────────────────────────────────────────────────────────
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:G{fila - 1}"

# Guardar
output = r"C:\cazalleida . com\cazadores-grayena-julio-2026.xlsx"
wb.save(output)
print(f"OK Excel guardado: {output}")
